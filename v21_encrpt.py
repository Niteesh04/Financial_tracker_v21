#!/usr/bin/env python3
"""
v21_encrpt.py — Secure Finance Tracker v21 (simplified schema)

Schema now only contains:
- Date (auto)
- Pocket Money (pocket)
- Extra Income (extra)
- Food & Drinks (food)
- Other Spending (other)
- Note (note) -- encrypted at field-level
- Tags (tags) -- encrypted at field-level

Features:
- SQLite DB + CSV + Excel sync
- PBKDF2-HMAC-SHA256 passphrase support (env FINANCE_PASSPHRASE) or fallback secret.key
- Field-level encryption (Fernet) for note & tags
- Encrypted artifacts saved on disk: CSV.enc, XLSX.enc, DB dump .enc, STATE.enc
- Backup rotation and encrypted backups
- Charts (daily bar and period line charts) and tag cloud
- Flask web UI: dashboard, add, charts, search, analytics, exports
- CLI helpers: decrypt artifact, restore DB
- run_app() exposed for run.py
"""

from __future__ import annotations

import os
import io
import sys
import json
import stat
import tarfile
import shutil
import sqlite3
import base64
import logging
from datetime import datetime
from typing import Tuple, Optional, List

import pandas as pd
import matplotlib.pyplot as plt
from wordcloud import WordCloud
from openpyxl import Workbook, load_workbook
from flask import Flask, render_template_string, request, redirect, send_file, url_for, abort
from markupsafe import Markup
from colorama import Fore, init as colorama_init

# cryptography
from cryptography.fernet import Fernet, InvalidToken
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives import hashes

# -----------------------------------------------------------------------------
# Initialization
# -----------------------------------------------------------------------------
colorama_init(autoreset=True)
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

# -----------------------------------------------------------------------------
# Configuration & Paths
# -----------------------------------------------------------------------------
VERSION_TAG = "v21"
ROOT = os.path.abspath(os.path.dirname(__file__))   # e.g. .../16082025
DATA_DIR = os.path.join(ROOT, "data_v21")           # central data folder
os.makedirs(DATA_DIR, exist_ok=True)

# Core files (plaintext CSV/Excel/DB for convenience; encrypted artifacts also maintained)
DB_FILE = os.path.join(DATA_DIR, "finance.db")
CSV_FILE = os.path.join(DATA_DIR, "daily_finance_tracker.csv")
EXCEL_FILE = os.path.join(DATA_DIR, "daily_finance_tracker.xlsx")
STATE_FILE = os.path.join(DATA_DIR, "finance_state.json")

# Encrypted artifacts (at-rest encrypted)
CSV_ENC = os.path.join(DATA_DIR, "daily_finance_tracker.csv.enc")
XLSX_ENC = os.path.join(DATA_DIR, "daily_finance_tracker.xlsx.enc")
DB_DUMP_ENC = os.path.join(DATA_DIR, "finance_dump.sql.enc")
STATE_ENC = os.path.join(DATA_DIR, "finance_state.json.enc")

# Key and KDF salt
SECRET_KEY_FILE = os.path.join(DATA_DIR, "secret.key")
KDF_SALT_FILE = os.path.join(DATA_DIR, "kdf_salt.bin")

# Charts and backups
CHARTS_FOLDER = os.path.join(DATA_DIR, f"all_charts_{VERSION_TAG}")
DAILY_FOLDER = os.path.join(CHARTS_FOLDER, "daily")
MONTHLY_FOLDER = os.path.join(CHARTS_FOLDER, "monthly")
YEARLY_FOLDER = os.path.join(CHARTS_FOLDER, "yearly")
TAGCLOUD_FOLDER = os.path.join(CHARTS_FOLDER, "tag_cloud")
BACKUP_FOLDER = os.path.join(DATA_DIR, "backup")
for p in [CHARTS_FOLDER, DAILY_FOLDER, MONTHLY_FOLDER, YEARLY_FOLDER, TAGCLOUD_FOLDER, BACKUP_FOLDER]:
    os.makedirs(p, exist_ok=True)

# App
app = Flask(__name__)

# Other settings
BACKUP_RETENTION = 20
DATE_FMT = "%Y-%m-%d"

# CSV schema (note: "Additional Income" and "Starting Balance" removed per request)
CSV_COLUMNS = [
    "Date",
    "Pocket Money",
    "Extra Income",
    "Total Income",
    "Food & Drinks",
    "Other Spending",
    "Total Spent",
    "Balance",
    "Note",
    "Tags"
]

# -----------------------------------------------------------------------------
# Utility helpers
# -----------------------------------------------------------------------------
def log(msg: str, color=Fore.GREEN):
    print(color + msg)

def safe_int(x, default=0):
    try:
        return int(float(x))
    except Exception:
        return default

def ensure_dirs():
    for d in [DATA_DIR, CHARTS_FOLDER, DAILY_FOLDER, MONTHLY_FOLDER, YEARLY_FOLDER, TAGCLOUD_FOLDER, BACKUP_FOLDER]:
        os.makedirs(d, exist_ok=True)

ensure_dirs()

# -----------------------------------------------------------------------------
# Crypto utilities (PBKDF2 passphrase or fallback file key)
# -----------------------------------------------------------------------------
def _write_file_secure(path: str, data: bytes):
    with open(path, "wb") as f:
        f.write(data)
    try:
        os.chmod(path, stat.S_IRUSR | stat.S_IWUSR)  # 0o600
    except Exception:
        pass

def _load_or_create_salt() -> bytes:
    if os.path.exists(KDF_SALT_FILE):
        return open(KDF_SALT_FILE, "rb").read()
    salt = os.urandom(16)
    _write_file_secure(KDF_SALT_FILE, salt)
    return salt

def _derive_key_from_passphrase(passphrase: str) -> bytes:
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=_load_or_create_salt(), iterations=390000)
    return base64.urlsafe_b64encode(kdf.derive(passphrase.encode("utf-8")))

def _load_or_create_filekey() -> bytes:
    if os.path.exists(SECRET_KEY_FILE):
        return open(SECRET_KEY_FILE, "rb").read()
    key = Fernet.generate_key()
    _write_file_secure(SECRET_KEY_FILE, key)
    return key

def get_cipher() -> Fernet:
    passphrase = os.environ.get("FINANCE_PASSPHRASE")
    if passphrase is not None and passphrase.strip() == "":
        passphrase = None
    if passphrase:
        key = _derive_key_from_passphrase(passphrase)
    else:
        key = _load_or_create_filekey()
    return Fernet(key)

CIPHER = get_cipher()

def enc_bytes(b: bytes) -> bytes:
    return CIPHER.encrypt(b)

def dec_bytes(b: bytes) -> bytes:
    return CIPHER.decrypt(b)

def enc_text(s: Optional[str]) -> str:
    if s is None or s == "":
        return ""
    return CIPHER.encrypt(s.encode("utf-8")).decode("utf-8")

def dec_text(s: Optional[str]) -> str:
    if s is None or s == "":
        return ""
    try:
        return CIPHER.decrypt(s.encode("utf-8")).decode("utf-8")
    except (InvalidToken, Exception):
        # fallback to plaintext if decryption fails
        return s

# -----------------------------------------------------------------------------
# Database initialization & migration
# -----------------------------------------------------------------------------
def create_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS finance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            pocket INTEGER,
            extra INTEGER,
            total_income INTEGER,
            food INTEGER,
            other INTEGER,
            total_spent INTEGER,
            balance INTEGER,
            note TEXT,
            tags TEXT,
            created_at TEXT
        )
    ''')
    # migration safety: try to add missing columns
    try:
        c.execute("ALTER TABLE finance ADD COLUMN total_income INTEGER DEFAULT 0")
    except Exception:
        pass
    try:
        c.execute("ALTER TABLE finance ADD COLUMN total_spent INTEGER DEFAULT 0")
    except Exception:
        pass
    try:
        c.execute("ALTER TABLE finance ADD COLUMN created_at TEXT")
    except Exception:
        pass
    conn.commit()
    conn.close()

create_db()

# -----------------------------------------------------------------------------
# State (plain JSON + encrypted copy)
# -----------------------------------------------------------------------------
def load_state() -> Tuple[int,int]:
    # try plaintext state
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                return int(data.get("day_index", 0)), int(data.get("balance_rollover", 0))
        except Exception:
            pass
    # try encrypted
    if os.path.exists(STATE_ENC):
        try:
            raw = dec_bytes(open(STATE_ENC, "rb").read())
            data = json.loads(raw.decode("utf-8"))
            return int(data.get("day_index", 0)), int(data.get("balance_rollover", 0))
        except Exception:
            pass
    return 0, 0

def save_state(day_index: int, balance_rollover: int):
    try:
        with open(STATE_FILE, "w", encoding="utf-8") as f:
            json.dump({"day_index": day_index, "balance_rollover": balance_rollover}, f)
    except Exception:
        pass
    blob = json.dumps({"day_index": day_index, "balance_rollover": balance_rollover}).encode("utf-8")
    _write_file_secure(STATE_ENC, enc_bytes(blob))

# -----------------------------------------------------------------------------
# CSV / Excel helpers (and encrypted copies)
# -----------------------------------------------------------------------------
def _append_to_csv(row_dict: dict):
    cols = CSV_COLUMNS
    df = pd.DataFrame([row_dict], columns=cols)
    if os.path.exists(CSV_FILE):
        try:
            current = pd.read_csv(CSV_FILE)
            missing = [c for c in cols if c not in current.columns]
            if missing:
                log("CSV missing columns. Rebuilding CSV from DB…", Fore.YELLOW)
                rebuild_csv_from_db()
            df.to_csv(CSV_FILE, mode='a', header=False, index=False)
        except Exception:
            rebuild_csv_from_db()
            df.to_csv(CSV_FILE, mode='a', header=False, index=False)
    else:
        df.to_csv(CSV_FILE, index=False)

def rebuild_csv_from_db():
    if not os.path.exists(DB_FILE):
        return
    conn = sqlite3.connect(DB_FILE)
    q = """SELECT
              date as "Date",
              pocket as "Pocket Money",
              extra as "Extra Income",
              total_income as "Total Income",
              food as "Food & Drinks",
              other as "Other Spending",
              total_spent as "Total Spent",
              balance as "Balance",
              note as "Note",
              tags as "Tags"
           FROM finance
           ORDER BY date ASC, id ASC"""
    df = pd.read_sql_query(q, conn)
    conn.close()
    # decrypt Note/Tags for CSV convenience
    try:
        if not df.empty:
            df["Note"] = df["Note"].astype(str).apply(lambda x: dec_text(x))
            df["Tags"] = df["Tags"].astype(str).apply(lambda x: dec_text(x))
    except Exception:
        pass
    df.to_csv(CSV_FILE, index=False)

def persist_csv_and_excel_encrypted():
    # Build DF
    if os.path.exists(CSV_FILE):
        try:
            df = pd.read_csv(CSV_FILE)
        except Exception:
            rebuild_csv_from_db()
            df = pd.read_csv(CSV_FILE) if os.path.exists(CSV_FILE) else pd.DataFrame(columns=CSV_COLUMNS)
    else:
        rebuild_csv_from_db()
        df = pd.read_csv(CSV_FILE) if os.path.exists(CSV_FILE) else pd.DataFrame(columns=CSV_COLUMNS)
    # Encrypted CSV
    csv_buf = io.StringIO()
    df.to_csv(csv_buf, index=False)
    _write_file_secure(CSV_ENC, enc_bytes(csv_buf.getvalue().encode("utf-8")))
    # Encrypted Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily_Records"
    ws.append(CSV_COLUMNS)
    for _, r in df.iterrows():
        ws.append([r.get(c, "") for c in CSV_COLUMNS])
    xbuf = io.BytesIO()
    wb.save(xbuf)
    _write_file_secure(XLSX_ENC, enc_bytes(xbuf.getvalue()))

# -----------------------------------------------------------------------------
# DB dump persistence (encrypted)
# -----------------------------------------------------------------------------
def persist_db_dump():
    conn = sqlite3.connect(DB_FILE)
    dump = "\n".join(conn.iterdump()).encode("utf-8")
    conn.close()
    _write_file_secure(DB_DUMP_ENC, enc_bytes(dump))

def restore_from_db_dump_if_any():
    if not os.path.exists(DB_DUMP_ENC):
        return
    try:
        sql = dec_bytes(open(DB_DUMP_ENC, "rb").read()).decode("utf-8", errors="ignore")
        conn = sqlite3.connect(DB_FILE)
        conn.executescript(sql)
        conn.commit()
        conn.close()
        log("Restored DB from encrypted dump.", Fore.CYAN)
    except Exception as e:
        log(f"DB restore failed (continuing with empty DB): {e}", Fore.YELLOW)

# -----------------------------------------------------------------------------
# Backups (encrypted)
# -----------------------------------------------------------------------------
def _rotate_backups(prefix: str):
    files = sorted([f for f in os.listdir(BACKUP_FOLDER) if f.startswith(prefix)], reverse=True)
    for f in files[BACKUP_RETENTION:]:
        try:
            os.remove(os.path.join(BACKUP_FOLDER, f))
        except Exception:
            pass

def backup_all():
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    # persist latest artifacts
    persist_db_dump()
    persist_csv_and_excel_encrypted()
    def copy_enc(src: str, tag: str):
        if os.path.exists(src):
            try:
                plain = dec_bytes(open(src, "rb").read())
            except Exception:
                with open(src, "rb") as f:
                    plain = f.read()
            out = os.path.join(BACKUP_FOLDER, f"{tag}_{ts}.enc")
            _write_file_secure(out, enc_bytes(plain))
            _rotate_backups(f"{tag}_")
    copy_enc(DB_DUMP_ENC, "dbdump")
    copy_enc(CSV_ENC, "csv")
    copy_enc(XLSX_ENC, "excel")
    if os.path.exists(STATE_ENC):
        copy_enc(STATE_ENC, "state")

# -----------------------------------------------------------------------------
# Tagging helper
# -----------------------------------------------------------------------------
def auto_tag(note_text: Optional[str]) -> str:
    s = (note_text or "").lower()
    tags = []
    if any(x in s for x in ["tired","lazy","rest","sleep"]): tags.append("#skipday")
    if any(x in s for x in ["short","half","partial"]): tags.append("#shortday")
    if any(x in s for x in ["save","no spend","zero spend"]): tags.append("#savings")
    if any(x in s for x in ["food","coffee","drink","meal"]): tags.append("#food")
    if any(x in s for x in ["uber","bus","trip","travel","metro","rickshaw"]): tags.append("#travel")
    seen = set(); out = []
    for t in tags:
        if t not in seen:
            seen.add(t); out.append(t)
    return " ".join(out)

# -----------------------------------------------------------------------------
# Charts
# -----------------------------------------------------------------------------
def generate_daily_chart(date: str, total_income: int, total_spent: int, balance: int):
    plt.figure(figsize=(7,4))
    labels = ['Total Income','Total Spent','Balance']
    values = [total_income, total_spent, balance]
    bars = plt.bar(labels, values)
    for b in bars:
        y = b.get_height()
        plt.text(b.get_x()+b.get_width()/2, y + max(1, 0.02*y), f"₹{int(y)}", ha='center')
    plt.title(f"Daily Summary {date}")
    plt.tight_layout()
    out = os.path.join(DAILY_FOLDER, f"{date}.png")
    plt.savefig(out); plt.close()
    return out

def generate_period_chart(df: pd.DataFrame, folder: str, period_name: str):
    if df.empty:
        return None
    plt.figure(figsize=(10,5))
    plt.plot(df['Date'], df['Total Income'], marker='o', label='Income')
    plt.plot(df['Date'], df['Total Spent'], marker='o', label='Spent')
    plt.plot(df['Date'], df['Balance'], marker='o', label='Balance')
    plt.xticks(rotation=45); plt.legend(); plt.grid(True)
    filename = os.path.join(folder, f"{period_name}_{df['Date'].iloc[0]}_to_{df['Date'].iloc[-1]}.png")
    plt.tight_layout(); plt.savefig(filename); plt.close()
    return filename

def generate_tag_cloud_from_series(series: pd.Series):
    text = " ".join([str(x) for x in series if isinstance(x, str)])
    if not text.strip(): return None
    wc = WordCloud(width=800, height=400, background_color='white').generate(text)
    out = os.path.join(TAGCLOUD_FOLDER, "tag_cloud.png")
    wc.to_file(out)
    return out

def optimized_chart_update(df: pd.DataFrame):
    if df.empty: return
    last = df.iloc[-1]
    generate_daily_chart(last['Date'], int(last['Total Income']), int(last['Total Spent']), int(last['Balance']))
    n = len(df)
    if n >= 30 and n % 30 == 0:
        generate_period_chart(df.tail(30), MONTHLY_FOLDER, "Monthly")
    if n >= 365 and n % 365 == 0:
        generate_period_chart(df.tail(365), YEARLY_FOLDER, "Yearly")
    tags = df['Tags'].fillna("")
    notes = df['Note'].fillna("")
    series = pd.Series([t if t else n for t,n in zip(tags, notes)])
    generate_tag_cloud_from_series(series)

# -----------------------------------------------------------------------------
# Save entry (DB, Excel, CSV). NOTE: simplified schema (no starting balance, no additional)
# -----------------------------------------------------------------------------
def save_entry(date: str, pocket: int, extra: int, new_total_income: int, food: int, other: int,
               total_spent: int, balance: int, note: str, tags: str):
    enc_note = enc_text(note)
    enc_tags = enc_text(tags)
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("""INSERT INTO finance
                 (date, pocket, extra, total_income, food, other, total_spent, balance, note, tags, created_at)
                 VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
              (date, pocket, extra, new_total_income, food, other, total_spent, balance, enc_note, enc_tags, datetime.now().isoformat()))
    conn.commit()
    conn.close()
    # Excel: ensure header exists
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily_Records"
        ws.append(CSV_COLUMNS)
        wb.save(EXCEL_FILE)
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Daily_Records"]
    ws.append([date, pocket, extra, new_total_income, food, other, total_spent, balance, note, tags])
    wb.save(EXCEL_FILE)
    # CSV
    _append_to_csv({
        "Date": date,
        "Pocket Money": pocket,
        "Extra Income": extra,
        "Total Income": new_total_income,
        "Food & Drinks": food,
        "Other Spending": other,
        "Total Spent": total_spent,
        "Balance": balance,
        "Note": note,
        "Tags": tags
    })
    # persist encrypted artifacts and backup
    persist_csv_and_excel_encrypted()
    persist_db_dump()
    backup_all()

# -----------------------------------------------------------------------------
# Load DataFrame from CSV or rebuild from DB
# -----------------------------------------------------------------------------
def load_df() -> pd.DataFrame:
    if os.path.exists(CSV_FILE):
        try:
            df = pd.read_csv(CSV_FILE)
            needed = CSV_COLUMNS
            for c in needed:
                if c not in df.columns:
                    rebuild_csv_from_db()
                    df = pd.read_csv(CSV_FILE)
                    break
            return df
        except Exception:
            rebuild_csv_from_db()
            if os.path.exists(CSV_FILE):
                return pd.read_csv(CSV_FILE)
            return pd.DataFrame(columns=CSV_COLUMNS)
    else:
        rebuild_csv_from_db()
        if os.path.exists(CSV_FILE):
            return pd.read_csv(CSV_FILE)
        return pd.DataFrame(columns=CSV_COLUMNS)

# -----------------------------------------------------------------------------
# Flask Routes
# -----------------------------------------------------------------------------
@app.route("/")
def dashboard():
    df = load_df()
    if df.empty:
        return redirect(url_for("add_entry"))
    view = df.tail(30).copy()
    df_html = view.to_html(classes="table table-striped", index=False)
    return render_template_string("""
    <html>
    <head>
      <title>Finance Dashboard ({{version}})</title>
      <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css">
    </head>
    <body class="p-4">
      <h1>Finance Dashboard — {{version}}</h1>
      <div class="mb-3">
        <a class="btn btn-success" href="{{ url_for('add_entry') }}">Add Entry</a>
        <a class="btn btn-secondary" href="{{ url_for('charts') }}">View Charts</a>
        <a class="btn btn-info" href="{{ url_for('search') }}">Search</a>
        <a class="btn btn-warning" href="{{ url_for('analytics') }}">Analytics</a>
        <a class="btn btn-outline-primary" href="{{ url_for('export_csv') }}">Download CSV</a>
        <a class="btn btn-outline-primary" href="{{ url_for('export_excel') }}">Download Excel</a>
        <a class="btn btn-outline-primary" href="{{ url_for('export_db') }}">Download DB Dump</a>
        <a class="btn btn-outline-danger" href="{{ url_for('export_secure') }}">Secure Export</a>
      </div>
      <hr>
      {{df_html | safe}}
    </body>
    </html>
    """, df_html=Markup(df_html), version=VERSION_TAG)

@app.route("/add", methods=["GET","POST"])
def add_entry():
    if request.method == "POST":
        today = datetime.now().strftime(DATE_FMT)
        pocket = safe_int(request.form.get("pocket") or 0)
        extra = safe_int(request.form.get("extra") or 0)
        food = safe_int(request.form.get("food") or 0)
        other = safe_int(request.form.get("other") or 0)
        note = (request.form.get("note") or "").strip()
        tags = auto_tag(note)
        total_income = pocket + extra
        total_spent = food + other
        balance = total_income - total_spent
        save_entry(today, pocket, extra, total_income, food, other, total_spent, balance, note, tags)
        # update charts (optimized)
        df = load_df()
        optimized_chart_update(df)
        return redirect(url_for("dashboard"))
    # GET form
    return render_template_string("""
    <html>
    <head>
      <title>Add Entry</title>
      <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css">
    </head>
    <body class="p-4">
      <h1>Add Finance Entry</h1>
      <form method="post" class="row g-3">
        <div class="col-md-4">
          <label class="form-label">Pocket Money</label>
          <input type="number" name="pocket" class="form-control" required>
        </div>
        <div class="col-md-4">
          <label class="form-label">Extra Income</label>
          <input type="number" name="extra" class="form-control" required>
        </div>
        <div class="col-md-4">
          <label class="form-label">Food & Drinks</label>
          <input type="number" name="food" class="form-control" value="0">
        </div>
        <div class="col-md-4">
          <label class="form-label">Other Spending</label>
          <input type="number" name="other" class="form-control" value="0">
        </div>
        <div class="col-12">
          <label class="form-label">Note</label>
          <textarea name="note" class="form-control" rows="2" placeholder="Write your note; tags auto-detected."></textarea>
        </div>
        <div class="col-12">
          <button type="submit" class="btn btn-success">Save</button>
          <a href="{{ url_for('dashboard') }}" class="btn btn-outline-secondary">Cancel</a>
        </div>
      </form>
    </body>
    </html>
    """)

@app.route("/charts")
def charts():
    files = []
    for folder in [DAILY_FOLDER, MONTHLY_FOLDER, YEARLY_FOLDER, TAGCLOUD_FOLDER]:
        if os.path.exists(folder):
            for f in sorted(os.listdir(folder)):
                files.append((folder, f))
    if not files:
        return "<h3>No charts available yet.</h3>"
    files_html = "".join([f"<li><a href='{url_for('serve_chart', folder=folder, filename=f)}'>{os.path.basename(folder)}/ {f}</a></li>" for folder, f in files])
    return f"<h1>Charts</h1><ul>{files_html}</ul>"

@app.route("/chart/<path:folder>/<path:filename>")
def serve_chart(folder, filename):
    abs_folder = os.path.abspath(folder)
    abs_base = os.path.abspath(CHARTS_FOLDER)
    if not abs_folder.startswith(abs_base):
        return "Access Denied."
    file_path = os.path.join(folder, filename)
    if os.path.exists(file_path):
        return send_file(file_path)
    return "File not found."

@app.route("/search", methods=["GET","POST"])
def search():
    if request.method == "POST":
        term = (request.form.get("term") or "").lower().strip()
        df = load_df()
        if df.empty:
            return "<h3>No data available to search.</h3>"
        result = df[df.apply(lambda row: row.astype(str).str.lower().str.contains(term).any(), axis=1)]
        if result.empty:
            return "<h3>No matching entries found.</h3>"
        return result.to_html(classes="table table-striped", index=False)
    return render_template_string("""
    <h1>Search Entries</h1>
    <form method="post">
      <input type="text" name="term" class="form-control" placeholder="Search term">
      <br>
      <button type="submit" class="btn btn-primary">Search</button>
    </form>
    """)

@app.route("/analytics")
def analytics():
    df = load_df()
    if df.empty:
        return "<h2>No data yet for analytics.</h2>"
    for col in ["Total Income", "Food & Drinks", "Other Spending", "Total Spent", "Balance"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
    avg_income = df["Total Income"].mean()
    avg_spent = df["Total Spent"].mean()
    avg_balance = df["Balance"].mean()
    total_days = len(df)
    total_spending = df["Total Spent"].sum()
    best_row = df.loc[df["Balance"].idxmax()] if total_days else None
    worst_row = df.loc[df["Balance"].idxmin()] if total_days else None
    tag_series = df["Tags"].dropna().astype(str)
    tags_flat = []
    for t in tag_series:
        tags_flat.extend([x for x in t.split() if x.startswith("#")])
    tag_counts = pd.Series(tags_flat).value_counts().head(10).to_dict() if tags_flat else {}
    df["Month"] = df["Date"].astype(str).str.slice(0, 7)
    by_month = df.groupby("Month").agg({"Total Income":"sum","Total Spent":"sum","Balance":"last"}).reset_index()
    by_month_html = by_month.to_html(classes="table table-bordered table-sm", index=False)
    return render_template_string("""
    <html>
    <head>
      <title>Analytics</title>
      <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css">
    </head>
    <body class="p-4">
      <h1>📊 Finance Analytics — {{version}}</h1>
      <a href="{{ url_for('dashboard') }}" class="btn btn-secondary mb-3">Back</a>
      <hr>
      <h3>Summary Stats</h3>
      <ul>
        <li>Average Daily Income: ₹{{avg_income|round(2)}}</li>
        <li>Average Daily Spending: ₹{{avg_spent|round(2)}}</li>
        <li>Average Daily Balance: ₹{{avg_balance|round(2)}}</li>
        <li>Total Days Recorded: {{total_days}}</li>
        <li>Total Spending Overall: ₹{{total_spending}}</li>
      </ul>
      <h3>Best / Worst Days (by Balance)</h3>
      <ul>
        <li><b>Best:</b> {{best_date}} (Balance ₹{{best_bal}})</li>
        <li><b>Worst:</b> {{worst_date}} (Balance ₹{{worst_bal}})</li>
      </ul>
      <h3>Top Tags</h3>
      <ul>
      {% for t, c in tag_counts.items() %}
        <li>{{t}} — {{c}} days</li>
      {% else %}
        <li>No tags yet.</li>
      {% endfor %}
      </ul>
      <h3>Monthly Overview</h3>
      {{by_month_html | safe}}
    </body>
    </html>
    """,
    version=VERSION_TAG,
    avg_income=avg_income,
    avg_spent=avg_spent,
    avg_balance=avg_balance,
    total_days=total_days,
    total_spending=total_spending,
    best_date=(best_row["Date"] if best_row is not None else "—"),
    best_bal=(int(best_row["Balance"]) if best_row is not None else 0),
    worst_date=(worst_row["Date"] if worst_row is not None else "—"),
    worst_bal=(int(worst_row["Balance"]) if worst_row is not None else 0),
    tag_counts=tag_counts,
    by_month_html=by_month_html
    )

# -----------------------------------------------------------------------------
# Exports
# -----------------------------------------------------------------------------
@app.route("/export/csv")
def export_csv():
    if os.path.exists(CSV_FILE):
        return send_file(CSV_FILE, as_attachment=True)
    return "CSV not found."

@app.route("/export/excel")
def export_excel():
    # ensure excel exists
    if not os.path.exists(EXCEL_FILE):
        if os.path.exists(CSV_FILE):
            try:
                pd.read_csv(CSV_FILE).to_excel(EXCEL_FILE, index=False)
            except Exception:
                rebuild_csv_from_db()
                pd.read_csv(CSV_FILE).to_excel(EXCEL_FILE, index=False)
        else:
            rebuild_csv_from_db()
            if os.path.exists(CSV_FILE):
                pd.read_csv(CSV_FILE).to_excel(EXCEL_FILE, index=False)
    if os.path.exists(EXCEL_FILE):
        return send_file(EXCEL_FILE, as_attachment=True)
    return "Excel not found."

@app.route("/export/db")
def export_db():
    persist_db_dump()
    if os.path.exists(DB_DUMP_ENC):
        return send_file(DB_DUMP_ENC, as_attachment=True)
    return "DB dump not found."

def secure_tarball_bytes(paths: List[str]) -> bytes:
    buf = io.BytesIO()
    with tarfile.open(fileobj=buf, mode="w:gz") as tar:
        for p in paths:
            if os.path.exists(p):
                arcname = os.path.relpath(p, start=DATA_DIR)
                tar.add(p, arcname=arcname)
    return enc_bytes(buf.getvalue())

@app.route("/export/secure")
def export_secure():
    try:
        persist_db_dump()
        persist_csv_and_excel_encrypted()
        ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        out_path = os.path.join(DATA_DIR, f"secure_export_{ts}.tar.gz.enc")
        blob = secure_tarball_bytes([DB_DUMP_ENC, CSV_ENC, XLSX_ENC, STATE_ENC])
        _write_file_secure(out_path, blob)
        return send_file(out_path, as_attachment=True)
    except Exception as e:
        abort(500, f"Secure export failed: {e}")

# -----------------------------------------------------------------------------
# CLI helpers
# -----------------------------------------------------------------------------
def decrypt_artifact(enc_path: str, out_path: str):
    with open(enc_path, "rb") as f:
        enc = f.read()
    plain = dec_bytes(enc)
    with open(out_path, "wb") as f:
        f.write(plain)
    return out_path

def restore_from_encrypted_dump(enc_path: str):
    tmp = os.path.join(DATA_DIR, "__tmp_dump.sql")
    decrypt_artifact(enc_path, tmp)
    with open(tmp, "r", encoding="utf-8", errors="ignore") as f:
        sql = f.read()
    conn = sqlite3.connect(DB_FILE)
    conn.executescript("DROP TABLE IF EXISTS finance;")
    conn.executescript(sql)
    conn.commit()
    conn.close()
    os.remove(tmp)
    rebuild_csv_from_db()
    persist_csv_and_excel_encrypted()
    return True

# -----------------------------------------------------------------------------
# Bootstrap
# -----------------------------------------------------------------------------
def bootstrap():
    create_db()
    restore_from_db_dump_if_any()
    if not os.path.exists(CSV_FILE):
        rebuild_csv_from_db()
    persist_csv_and_excel_encrypted()
    log(f"Bootstrap complete — data dir: {DATA_DIR}", Fore.CYAN)

bootstrap()

# -----------------------------------------------------------------------------
# Expose run_app
# -----------------------------------------------------------------------------
def run_app(host="127.0.0.1", port=5000, debug=False):
    app.run(host=host, port=port, debug=debug)

# Allow CLI use
if __name__ == "__main__":
    if len(sys.argv) > 1:
        cmd = sys.argv[1].lower()
        if cmd == "decrypt" and len(sys.argv) >= 3:
            decrypt_artifact(sys.argv[2], sys.argv[3] if len(sys.argv) >= 4 else sys.argv[2] + ".dec")
            print("Decrypted.")
            sys.exit(0)
        if cmd == "restore" and len(sys.argv) >= 2:
            restore_from_encrypted_dump(sys.argv[2])
            print("Restored from dump.")
            sys.exit(0)
    print(f"Starting v21 web app (data dir: {DATA_DIR})")
    run_app()
